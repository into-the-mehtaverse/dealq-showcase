# Utility functions for Excel model registration

import json
import aiofiles
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
from fastapi import HTTPException, UploadFile
import openpyxl
from openpyxl.utils import get_column_letter
from langchain_openai import ChatOpenAI
from langchain.schema import HumanMessage

def parse_llm_json_response(response, context: str = "LLM call") -> Dict[str, Any]:
    """
    Helper function to parse JSON from LLM responses, handling markdown code blocks.

    Args:
        response: LangChain response object
        context: Description of what the LLM call was for (for error logging)

    Returns:
        Parsed JSON dictionary

    Raises:
        ValueError: If parsing fails
    """
    try:
        # Extract response content
        response_text = response.content if hasattr(response, 'content') else str(response)

        if not response_text or response_text.strip() == "":
            raise ValueError(f"Empty response from LLM in {context}")

        # Strip markdown code blocks (common with GPT-4o)
        response_text = response_text.strip()
        if response_text.startswith("```json"):
            response_text = response_text[7:]  # Remove ```json
        elif response_text.startswith("```"):
            response_text = response_text[3:]   # Remove ```

        if response_text.endswith("```"):
            response_text = response_text[:-3]  # Remove trailing ```

        response_text = response_text.strip()

        # Parse JSON
        result = json.loads(response_text)
        return result

    except json.JSONDecodeError as e:
        print(f"JSON decode error in {context}: {str(e)}")
        print(f"Attempted to parse: '{response_text[:500] if 'response_text' in locals() else 'N/A'}'")
        raise ValueError(f"Failed to parse JSON in {context}: {str(e)}")
    except Exception as e:
        print(f"General error parsing LLM response in {context}: {str(e)}")
        raise ValueError(f"Error processing LLM response in {context}: {str(e)}")

# Directory setup
BASE_DIR = Path(__file__).parent.parent.parent.parent  # Go up to project root
MODELS_DIR = BASE_DIR / "files" / "uploads" / "models"
MAPPINGS_DIR = BASE_DIR / "files" / "outputs" / "mappings"
METADATA_DIR = MODELS_DIR / "metadata"

# Ensure directories exist
MODELS_DIR.mkdir(parents=True, exist_ok=True)
MAPPINGS_DIR.mkdir(parents=True, exist_ok=True)
METADATA_DIR.mkdir(parents=True, exist_ok=True)

# Export directory constants for use in other modules
__all__ = [
    'MODELS_DIR', 'MAPPINGS_DIR', 'METADATA_DIR',
    'validate_excel_file', 'save_excel_model', 'analyze_excel_structure',
    'verify_multifamily_model', 'identify_input_sheets', 'generate_input_mappings',
    'save_model_mapping', 'create_model_metadata'
]

def validate_excel_file(file: UploadFile) -> None:
    """Validate that the uploaded file is an Excel file."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    valid_extensions = ['.xlsx', '.xls', '.xlsm']
    if not any(file.filename.lower().endswith(ext) for ext in valid_extensions):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls, .xlsm) are allowed")


async def save_excel_model(file: UploadFile, model_id: str) -> str:
    """
    Save an uploaded Excel file to the models directory.
    Returns the saved filename.
    """
    # Preserve the original file extension
    original_extension = Path(file.filename).suffix.lower()
    saved_filename = f"{model_id}{original_extension}"
    file_path = MODELS_DIR / saved_filename

    try:
        # Save file asynchronously
        async with aiofiles.open(file_path, 'wb') as f:
            content = await file.read()
            await f.write(content)

        return saved_filename
    except Exception as e:
        # Clean up partial file if save failed
        if file_path.exists():
            file_path.unlink()
        raise HTTPException(status_code=500, detail=f"Failed to save Excel file: {str(e)}")

def analyze_excel_structure(workbook_path: Path) -> Dict[str, Any]:
    """
    Analyze Excel file structure using openpyxl to extract technical information.

    Returns:
        Dictionary containing sheet information, cell types, formulas, etc.
    """
    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=False)

        structure = {
            "filename": workbook_path.name,
            "sheet_names": workbook.sheetnames,
            "sheets": {}
        }

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column

            # Analyze cell contents
            formulas = []
            input_cells = []
            headers = []
            data_validation = []

            # Sample first 20 rows for analysis (to avoid processing huge sheets)
            sample_rows = min(20, max_row)

            for row in range(1, sample_rows + 1):
                for col in range(1, min(26, max_col + 1)):  # First 26 columns (A-Z)
                    cell = sheet.cell(row=row, column=col)
                    cell_ref = f"{get_column_letter(col)}{row}"

                    if cell.value is not None:
                        # Check if cell contains a formula
                        if str(cell.value).startswith('='):
                            formulas.append({
                                "cell": cell_ref,
                                "formula": str(cell.value)
                            })
                        else:
                            # Potential input cell or header
                            cell_info = {
                                "cell": cell_ref,
                                "value": str(cell.value),
                                "data_type": str(type(cell.value).__name__)
                            }

                            # Check if it looks like a header (row 1-3, contains text)
                            if row <= 3 and isinstance(cell.value, str):
                                headers.append(cell_info)
                            else:
                                input_cells.append(cell_info)

                    # Note: Data validation is checked at worksheet level, not cell level

            # Collect data validation rules at worksheet level
            try:
                for dv in sheet.data_validations.dataValidation:
                    for cell_range in dv.cells:
                        data_validation.append({
                            "range": str(cell_range),
                            "type": str(dv.type) if dv.type else None,
                            "formula1": str(dv.formula1) if dv.formula1 else None,
                            "formula2": str(dv.formula2) if dv.formula2 else None
                        })
            except Exception:
                # If data validation collection fails, continue without it
                pass

            structure["sheets"][sheet_name] = {
                "dimensions": {"max_row": max_row, "max_col": max_col},
                "formulas": formulas,
                "headers": headers,
                "input_cells": input_cells[:100],  # Limit to first 100 for analysis
                "data_validation": data_validation,
                "formula_count": len(formulas),
                "input_cell_count": len(input_cells)
            }

        workbook.close()
        return structure

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to analyze Excel structure: {str(e)}")

async def verify_multifamily_model(excel_structure: Dict[str, Any], llm: ChatOpenAI) -> Dict[str, Any]:
    """
    Use LLM to verify if the Excel file is a multifamily real estate model.

    Returns:
        Dictionary with verification result and confidence
    """
    # Create a summary of the Excel structure for LLM analysis
    structure_summary = {
        "sheet_names": excel_structure["sheet_names"],
        "headers_by_sheet": {},
        "formula_patterns": {}
    }

    for sheet_name, sheet_data in excel_structure["sheets"].items():
        # Extract meaningful headers
        headers = [h["value"] for h in sheet_data["headers"] if len(h["value"]) > 2]
        structure_summary["headers_by_sheet"][sheet_name] = headers[:20]  # First 20 headers

        # Extract formula patterns (without actual formulas for privacy)
        formula_count = sheet_data["formula_count"]
        input_count = sheet_data["input_cell_count"]
        structure_summary["formula_patterns"][sheet_name] = {
            "formula_count": formula_count,
            "input_count": input_count,
            "formula_ratio": formula_count / (formula_count + input_count) if (formula_count + input_count) > 0 else 0
        }

    prompt = f"""
Analyze this Excel file structure to determine if it's a multifamily real estate financial model.

Excel Structure:
{json.dumps(structure_summary, indent=2)}

A multifamily real estate model typically contains:
- Rent roll analysis (unit types, rents, occupancy)
- Operating statements (T-12 financials)

- Cash flow analysis
- Property valuation metrics
- Deal summary information

Analyze the sheet names and headers to determine:
1. Is this a multifamily real estate financial model?
2. What confidence level (high/medium/low)?
3. What evidence supports your conclusion?

Respond in JSON format:
{{
    "is_multifamily_model": true/false,
    "confidence": "high/medium/low",
    "reason": "detailed explanation",
    "evidence": ["list of supporting evidence"],
    "missing_elements": ["list of typical elements not found"]
}}
"""

    try:
        response = await llm.ainvoke([HumanMessage(content=prompt)])

        # Use helper function to parse JSON response
        result = parse_llm_json_response(response, "model verification")

        # Validate response structure
        required_keys = ["is_multifamily_model", "confidence", "reason"]
        if not all(key in result for key in required_keys):
            raise ValueError("Invalid LLM response format - missing required keys")

        return result

    except Exception as e:
        print(f"Error in model verification: {str(e)}")
        # Fallback response if verification fails
        return {
            "is_multifamily_model": False,
            "confidence": "low",
            "reason": f"LLM verification failed: {str(e)}",
            "evidence": [],
            "missing_elements": ["Unable to analyze due to error"]
        }

async def identify_input_sheets(excel_structure: Dict[str, Any], llm: ChatOpenAI) -> Dict[str, str]:
    """
    Use LLM to identify THE single input sheet for each category (rent_roll, t12).

    Returns:
        Dictionary mapping categories to their input sheet names
    """
    # Collect all sheet information for analysis
    sheets_summary = []
    for sheet_name, sheet_data in excel_structure["sheets"].items():
        headers = [h["value"] for h in sheet_data["headers"] if len(h["value"]) > 2]
        sheets_summary.append({
            "sheet_name": sheet_name,
            "headers": headers[:10],  # First 10 meaningful headers
            "formula_count": sheet_data["formula_count"],
            "input_cell_count": sheet_data["input_cell_count"]
        })

    prompt = f"""
Analyze these Excel sheets to identify THE single input sheet for each category. Real estate models have only ONE input location per category.

Sheets: {json.dumps(sheets_summary, indent=2)}

For each category, identify the ONE sheet where users input data (not calculated results):
- rent_roll: Where users enter unit details, rents, lease info
- t12: Where users enter historical financial line items


You must pick exactly ONE sheet per category or "none" if not found.

Respond in JSON format:
{{
    "rent_roll": "sheet_name_or_none",
    "t12": "sheet_name_or_none",

}}
"""

    try:
        response = await llm.ainvoke([HumanMessage(content=prompt)])

        # Use helper function to parse JSON response
        result = parse_llm_json_response(response, "input sheet identification")

        # Clean up response - only keep categories with actual sheet names
        cleaned_result = {}
        for category in ["rent_roll", "t12"]:
            sheet_name = result.get(category, "none")
            if sheet_name and sheet_name != "none":
                cleaned_result[category] = sheet_name

        return cleaned_result

    except Exception as e:
        print(f"Error identifying input sheets: {str(e)}")
        return {}

async def generate_input_mappings(
    workbook_path: Path,
    input_sheets: Dict[str, str],
    llm: ChatOpenAI
) -> Dict[str, Any]:
    """
    Generate tactical mappings for inserting structured data.

    Returns:
        Clean mappings focused only on data insertion
    """
    mappings = {}

    # Load workbook for detailed analysis (preserving VBA)
    workbook = openpyxl.load_workbook(workbook_path, data_only=False, keep_vba=True)

    # Generate mapping for each identified input sheet (tabular data)
    for category, sheet_name in input_sheets.items():
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            mapping = await generate_tactical_mapping(sheet, sheet_name, category, llm)

            if mapping:
                mappings[category] = mapping

    # Generate property info mapping (single-cell data across all sheets)
    property_mapping = await generate_property_info_mapping(workbook, llm)
    if property_mapping:
        mappings["property_info"] = property_mapping

    workbook.close()
    return mappings

async def generate_property_info_mapping(
    workbook,
    llm: ChatOpenAI
) -> Optional[Dict[str, Any]]:
    """Generate mapping for single-cell property information fields across all sheets."""

    print(f"🔍 Starting property info mapping across all sheets")

    try:
        # Extract cell samples from all sheets (focus on input areas, avoid heavy formula sheets)
        all_cells_sample = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Sample cells from first 26 rows and 26 columns to find property info
            for row in range(1, min(27, sheet.max_row + 1)):
                for col in range(1, min(27, sheet.max_column + 1)):
                    try:
                        cell = sheet.cell(row=row, column=col)
                        if cell.value and not str(cell.value).startswith('='):
                            cell_value = str(cell.value).strip()
                            # Only include meaningful content (avoid empty strings, single chars)
                            if len(cell_value) > 1:
                                all_cells_sample.append({
                                    "sheet": sheet_name,
                                    "cell": f"{get_column_letter(col)}{row}",
                                    "value": cell_value[:100]  # Truncate long values
                                })
                    except Exception:
                        continue

        print(f"   Extracted {len(all_cells_sample)} cell samples from all sheets")

        if len(all_cells_sample) == 0:
            print(f"   ❌ No cell content found - skipping property info mapping")
            return None

        # Define the property info fields we need to find
        property_fields = ["name", "address", "zip_code", "year_built", "parking_spots", "gross_square_feet", "broker_price"]

        prompt = f"""
Analyze this Excel workbook to find the exact cells where property information should be entered.

Cell Samples: {json.dumps(all_cells_sample[:100], indent=2)}

TASK: Find the exact cell location for each property field:
- name: Property name (usually near the top, may be labeled "Property Name", "Asset Name", etc.)
- address: Property address (street address)
- zip_code: ZIP/postal code
- year_built: Year the property was built/constructed
- parking_spots: Total number of parking spaces/spots
- gross_square_feet: Total building square footage (GSF, GLA, etc.)
- broker_price: Broker price or sale price of the property (may be labeled "Broker Price", "Asking Price", "Sale Price", "List Price", etc.)

RULES:
1. Find the exact cell where USER INPUT should go (not labels/headers)
2. Each field should map to exactly ONE cell
3. Look for cells that are typically filled in by users, not calculated
4. Skip cells that contain formulas or are obviously labels
5. If a field isn't found, omit it from the response

Respond with ONLY this JSON structure:
{{
    "type": "property_info",
    "cells": {{
        "name": {{"sheet": "SheetName", "cell": "B5"}},
        "address": {{"sheet": "SheetName", "cell": "B6"}},
        "zip_code": {{"sheet": "SheetName", "cell": "B7"}},
        "year_built": {{"sheet": "SheetName", "cell": "B8"}},
        "parking_spots": {{"sheet": "SheetName", "cell": "B9"}},
        "gross_square_feet": {{"sheet": "SheetName", "cell": "B10"}},
        "broker_price": {{"sheet": "SheetName", "cell": "B11"}}
    }}
}}
"""

        print(f"   🤖 Calling LLM for property info mapping...")
        response = await llm.ainvoke([HumanMessage(content=prompt)])
        print(f"   ✅ LLM responded for property info")

        # Use helper function to parse JSON response
        result = parse_llm_json_response(response, "property info mapping")
        print(f"   📋 Parsed result for property info: {result}")

        # Validate that we have the required structure
        if "type" in result and "cells" in result and result["type"] == "property_info":
            print(f"   ✅ Valid property info mapping structure")
            return result
        else:
            print(f"   ❌ Invalid property info mapping structure")
            return None

    except Exception as e:
        print(f"❌ Error generating property info mapping: {str(e)}")
        import traceback
        print(f"   Full traceback: {traceback.format_exc()}")
        return None


async def generate_tactical_mapping(
    sheet,
    sheet_name: str,
    category: str,
    llm: ChatOpenAI
) -> Optional[Dict[str, Any]]:
    """Generate focused mapping for tabular data insertion (rent_roll, t12)."""

    print(f"🔍 Starting tactical mapping for {category} on sheet '{sheet_name}'")

    try:
        # Get sheet dimensions for debugging
        print(f"   Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")

        # Extract only the essential header information (first 10 rows, first 15 columns)
        headers_sample = []
        for row in range(1, min(11, sheet.max_row + 1)):
            for col in range(1, min(16, sheet.max_column + 1)):
                try:
                    cell = sheet.cell(row=row, column=col)
                    if cell.value and not str(cell.value).startswith('='):
                        headers_sample.append({
                            "row": row,
                            "column": get_column_letter(col),
                            "value": str(cell.value)[:50]  # Truncate long values
                        })
                except Exception as cell_error:
                    print(f"   Warning: Error reading cell {get_column_letter(col)}{row}: {str(cell_error)}")
                    continue

        print(f"   Extracted {len(headers_sample)} header samples")
        if len(headers_sample) == 0:
            print(f"   ❌ No headers found in sheet '{sheet_name}' - skipping mapping")
            return None

        print(f"   First few headers: {headers_sample[:5]}")

        # Import field definitions from central mapping models
        from app.models.mapping import FIELD_DEFINITIONS
        field_definitions = FIELD_DEFINITIONS

        required_fields = field_definitions.get(category, [])
        print(f"   Required fields for {category}: {required_fields}")

        prompt = f"""
Analyze this Excel sheet to create a TACTICAL mapping for inserting {category} data.

Sheet: {sheet_name}
Sample Headers/Values: {json.dumps(headers_sample[:30], indent=2)}

TASK: Map these exact fields to their corresponding columns:
{required_fields}

RULES:
1. Find the row where data should start being inserted (usually after headers)
2. Map each field to its column letter (A, B, C, etc.)
3. Only include columns where we should INSERT data (skip calculated/formula columns)

Respond with ONLY this JSON structure:
{{
    "sheet_name": "{sheet_name}",
    "data_start_row": <number>,
    "columns": {{
        "field1": "A",
        "field2": "C",
        "field3": "E"
    }}
}}
"""

        print(f"   🤖 Calling LLM for {category} tactical mapping...")
        response = await llm.ainvoke([HumanMessage(content=prompt)])
        print(f"   ✅ LLM responded for {category}")

        # Use helper function to parse JSON response
        result = parse_llm_json_response(response, f"tactical mapping for {category}")
        print(f"   📋 Parsed result for {category}: {result}")

        # Validate that we have the required structure
        if "sheet_name" in result and "data_start_row" in result and "columns" in result:
            print(f"   ✅ Valid mapping structure for {category}")
            return result
        else:
            print(f"   ❌ Invalid mapping structure for {category} - missing required keys")
            print(f"   Required: sheet_name, data_start_row, columns")
            print(f"   Got: {list(result.keys()) if result else 'None'}")
            return None

    except Exception as e:
        print(f"❌ Error generating tactical mapping for {category} on sheet '{sheet_name}': {str(e)}")
        import traceback
        print(f"   Full traceback: {traceback.format_exc()}")
        return None

async def save_model_mapping(model_id: str, mapping: Dict[str, Any]) -> None:
    """Save the generated mapping to a JSON file."""
    mapping_path = MAPPINGS_DIR / f"{model_id}_mapping.json"

    try:
        async with aiofiles.open(mapping_path, 'w') as f:
            await f.write(json.dumps(mapping, indent=2))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save mapping: {str(e)}")

async def create_model_metadata(model_id: str, model_name: str, original_filename: str) -> None:
    """Create metadata file for the registered model."""
    metadata = {
        "model_id": model_id,
        "model_name": model_name,
        "original_filename": original_filename,
        "registered_at": datetime.now().isoformat()
    }

    metadata_path = METADATA_DIR / f"{model_id}.json"

    try:
        async with aiofiles.open(metadata_path, 'w') as f:
            await f.write(json.dumps(metadata, indent=2))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create metadata: {str(e)}")
