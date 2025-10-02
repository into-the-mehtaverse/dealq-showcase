"""
Rent Roll Extraction Service

Handles raw data extraction from PDF and Excel files for rent roll documents.
This service only extracts raw data - no structuring is performed.
The PDF/Excel file will be extracted in entirety first (extract_rent_roll_pdf/extract_rent_roll_excel),
then it will be passed through the precision extraction to extract only the relevant data (extract_rent_roll_with_boundaries).
"""

import os
import logging
import fitz  # PyMuPDF
import openpyxl
from typing import List, Dict, Any, Optional, Union
from fastapi import HTTPException
from .utils import (
    validate_file_path,
    validate_file_type,
    clean_excel_data,
    extract_text_from_pdf_page,
    create_extraction_metadata,
    convert_excel_data_to_enumerated_text,
    is_xml_security_error,
    log_file_safety_status,
    iter_rows_efficiently
)
from app.models.domain.rr_classification import (
    PDFRentRollClassification,
    ExcelRentRollClassification,
    PDFRentRollPrecisionExtract,
    ExcelRentRollPrecisionExtract
)

# Configure logging
logger = logging.getLogger(__name__)


class RentRollExtractionService:
    """Service for extracting raw data from PDF and Excel files."""

    def __init__(self):
        pass

    async def extract_rent_roll_pdf(self, file_path: str) -> Dict[str, Any]:
        """
        Extract raw rent roll data from PDF file.
        Args:
            file_path: Path to the PDF file
        Returns:
            Dictionary containing raw extracted rent roll data
        """
        try:
            # Validate file path and type
            validate_file_path(file_path)
            validate_file_type(file_path, "pdf")

            # Open the PDF document
            doc = fitz.open(file_path)
            extracted_data = {
                "file_type": "pdf",
                "document_type": "rent_roll",
                "total_pages": len(doc),
                "extracted_text": [],
                "page_data": []
            }

            # Extract text from each page
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                text = extract_text_from_pdf_page(page)

                page_data = {
                    "page_number": page_num + 1,
                    "text": text,
                    "text_length": len(text)
                }

                extracted_data["extracted_text"].append(text)
                extracted_data["page_data"].append(page_data)

            doc.close()

            return extracted_data

        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to extract rent roll PDF data: {str(e)}"
            )

    async def extract_rent_roll_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Extract raw rent roll data from Excel file.
        Args:
            file_path: Path to the Excel file
        Returns:
            Dictionary containing raw extracted rent roll data
        """
        try:
            print(f"    Starting RR Excel extraction: {file_path}")
            logger.info(f"Starting rent roll Excel extraction: {file_path}")

            # Validate file path and type
            validate_file_path(file_path)
            validate_file_type(file_path, "excel")

            # Load the Excel workbook
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True, keep_links=False)
                log_file_safety_status(file_path, is_safe=True)
                print(f"    Excel workbook loaded: {len(workbook.sheetnames)} sheets")
            except Exception as excel_error:
                # Check if the error is XML-related (indicating potential security issues)
                if is_xml_security_error(excel_error):
                    log_file_safety_status(file_path, is_safe=False, error=excel_error)
                    raise HTTPException(
                        status_code=400,
                        detail=f"Excel file appears to be unsafe or corrupted: {str(excel_error)}"
                    )
                else:
                    # Non-XML related error, re-raise
                    logger.error(f"Excel file parsing failed (non-XML error): {file_path}. Error: {excel_error}")
                    raise excel_error

            extracted_data = {
                "file_type": "excel",
                "document_type": "rent_roll",
                "sheets": [],
                "total_sheets": len(workbook.sheetnames)
            }

            # Extract data from each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                print(f"    Processing sheet: {sheet_name}")

                # Get all data from the sheet using memory-efficient iteration
                sheet_data = iter_rows_efficiently(sheet, max_consecutive_empty_rows=20)
                print(f"    Sheet {sheet_name}: {len(sheet_data)} rows after efficient processing")

                # Clean the sheet data
                cleaned_sheet_data = clean_excel_data(sheet_data)
                print(f"    Sheet {sheet_name}: {len(cleaned_sheet_data)} rows after cleaning")

                sheet_info = {
                    "sheet_name": sheet_name,
                    "data": cleaned_sheet_data,
                    "rows": len(cleaned_sheet_data),
                    "columns": len(cleaned_sheet_data[0]) if cleaned_sheet_data else 0
                }

                extracted_data["sheets"].append(sheet_info)

            workbook.close()
            print(f"    RR Excel extraction completed: {len(extracted_data['sheets'])} sheets")
            logger.info(f"Rent roll Excel extraction completed successfully: {file_path}")

            # Convert all sheets to enumerated text
            enumerated_text = convert_excel_data_to_enumerated_text(extracted_data["sheets"])
            extracted_data["enumerated_text"] = enumerated_text
            print(f"    RR Excel: Converted to {len(enumerated_text)} characters of enumerated text")

            return extracted_data

        except HTTPException:
            # Re-raise HTTP exceptions as-is
            raise
        except Exception as e:
            print(f"    RR Excel extraction error: {str(e)}")
            logger.error(f"Rent roll Excel extraction failed: {file_path}. Error: {str(e)}")
            raise HTTPException(
                status_code=500,
                detail=f"Failed to extract rent roll Excel data: {str(e)}"
            )

    def extract_rent_roll_with_boundaries(
        self,
        full_extraction: Dict[str, Any],
        classification_result: Union[PDFRentRollClassification, ExcelRentRollClassification]
    ) -> Union[PDFRentRollPrecisionExtract, ExcelRentRollPrecisionExtract]:
        """
        Extract only the relevant rent roll data based on classification boundaries.
        Args:
            full_extraction: Complete extracted data from extraction service
            classification_result: Classification result with boundaries
        Returns:
            Pydantic model containing only the relevant data within boundaries
        """
        file_type = full_extraction.get("file_type")

        if file_type == "pdf":
            return self._extract_pdf_with_boundaries(full_extraction, classification_result)
        elif file_type == "excel":
            return self._extract_excel_with_boundaries(full_extraction, classification_result)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    def _extract_pdf_with_boundaries(
        self,
        full_extraction: Dict[str, Any],
        classification: PDFRentRollClassification
    ) -> PDFRentRollPrecisionExtract:
        """Extract PDF data within specified page boundaries and content markers."""

        # Get page data
        page_data = full_extraction.get("page_data", [])

        # Filter pages within boundaries (inclusive)
        relevant_pages = []
        for page in page_data:
            page_num = page.get("page_number", 0)
            if classification.data_start_page <= page_num <= classification.data_end_page:
                relevant_pages.append(page)

        # Extract relevant content from each page using markers
        relevant_data = []
        for page in relevant_pages:
            text = page.get("text", "")
            page_content = []

            # Apply start marker if specified
            if classification.data_start_marker and classification.data_start_marker in text:
                start_idx = text.find(classification.data_start_marker)
                text = text[start_idx:]

            # Apply end marker if specified (inclusive)
            if classification.data_end_marker and classification.data_end_marker in text:
                end_idx = text.find(classification.data_end_marker)
                # Include the end marker by finding the end of the line containing it
                end_line_start = text.rfind('\n', 0, end_idx)
                if end_line_start == -1:
                    end_line_start = 0
                text = text[:end_idx + len(classification.data_end_marker)]

            # Split into lines and filter out empty lines
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            if lines:
                page_content.extend(lines)

            # Only add page if it has content
            if page_content:
                relevant_data.append(page_content)

        return PDFRentRollPrecisionExtract(
            total_pages=len(relevant_data),
            relevant_data=relevant_data,
            boundaries={
                "start_page": classification.data_start_page,
                "end_page": classification.data_end_page,
                "start_marker": classification.data_start_marker,
                "end_marker": classification.data_end_marker,
                "column_headers": classification.column_headers
            }
        )

    def _extract_excel_with_boundaries(
        self,
        full_extraction: Dict[str, Any],
        classification: ExcelRentRollClassification
    ) -> ExcelRentRollPrecisionExtract:
        """Extract Excel data within specified row boundaries."""

        # Get sheets data
        sheets = full_extraction.get("sheets", [])

        # Find the target sheet
        target_sheet = None
        for sheet in sheets:
            if sheet.get("sheet_name") == classification.data_sheet_name:
                target_sheet = sheet
                break

        if not target_sheet:
            raise ValueError(f"Sheet '{classification.data_sheet_name}' not found in extracted data")

        # Get data rows
        data = target_sheet.get("data", [])

        # Extract rows within boundaries (inclusive)
        start_row = classification.data_start_row - 1  # Convert to 0-based index
        end_row = classification.data_end_row  # Keep as 1-based for inclusive

        if start_row < 0:
            start_row = 0
        if end_row > len(data):
            end_row = len(data)

        relevant_data = data[start_row:end_row]

        # Create new sheet info with relevant data
        relevant_sheet = {
            "sheet_name": target_sheet.get("sheet_name"),
            "data": relevant_data,
            "rows": len(relevant_data),
            "columns": len(relevant_data[0]) if relevant_data else 0
        }

        # Convert to enumerated text for consistency
        enumerated_text = convert_excel_data_to_enumerated_text([relevant_sheet])

        return ExcelRentRollPrecisionExtract(
            sheets=[relevant_sheet],
            total_sheets=1,
            enumerated_text=enumerated_text,
            boundaries={
                "sheet_name": classification.data_sheet_name,
                "start_row": classification.data_start_row,
                "end_row": classification.data_end_row,
                "start_column": classification.data_start_column,
                "end_column": classification.data_end_column,
                "column_headers": classification.column_headers
            }
        )
