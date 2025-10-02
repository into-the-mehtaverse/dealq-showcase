"""
Extraction Service

Handles raw data extraction from PDF and Excel files for T12 documents.
This service only extracts raw data - no structuring is performed.
"""

import os
import logging
import fitz  # PyMuPDF
import openpyxl
from typing import List, Dict, Any, Optional, Union
from fastapi import HTTPException
from .utils import (
    validate_file_path,
    validate_file_type,
    clean_excel_data,
    extract_text_from_pdf_page,
    convert_excel_data_to_enumerated_text,
    iter_rows_efficiently
)

# Configure logging
logger = logging.getLogger(__name__)

class T12ExtractionService:
    """Service for extracting raw data from PDF and Excel files."""

    def __init__(self):
        pass

    async def extract_t12_pdf(self, file_path: str) -> Dict[str, Any]:
        """
        Extract T12 data from PDF file.
        Args:
            file_path: Path to the PDF file
        Returns:
            Dictionary containing raw extracted T12 data
        """
        try:
            # Validate file path and type
            validate_file_path(file_path)
            validate_file_type(file_path, "pdf")

            # Open the PDF document
            doc = fitz.open(file_path)
            extracted_data = {
                "file_type": "pdf",
                "document_type": "t12",
                "total_pages": len(doc),
                "extracted_text": [],
                "page_data": []
            }

            # Extract text from each page
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                text = extract_text_from_pdf_page(page)

                page_data = {
                    "page_number": page_num + 1,
                    "text": text,
                    "text_length": len(text)
                }

                extracted_data["extracted_text"].append(text)
                extracted_data["page_data"].append(page_data)

            doc.close()

            return extracted_data

        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to extract T12 PDF data: {str(e)}"
            )

    async def extract_t12_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Extract T12 data from Excel file.
        Args:
            file_path: Path to the Excel file
        Returns:
            Dictionary containing raw extracted T12 data
        """
        try:
            print(f"    Starting T12 Excel extraction: {file_path}")

            # Validate file path and type
            validate_file_path(file_path)
            validate_file_type(file_path, "excel")

            # Load the Excel workbook
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True, keep_links=False)
                logger.info(f"T12 Excel file parsed successfully - File is SAFE: {file_path}")
                print(f"    Excel workbook loaded: {len(workbook.sheetnames)} sheets")
            except Exception as excel_error:
                # Check if the error is XML-related (indicating potential security issues)
                error_str = str(excel_error).lower()
                if any(xml_indicator in error_str for xml_indicator in ['xml', 'bomb', 'external', 'entity', 'dtd']):
                    logger.warning(f"T12 Excel file may be UNSAFE - XML parsing error: {file_path}. Error: {excel_error}")
                    raise HTTPException(
                        status_code=400,
                        detail=f"T12 Excel file appears to be unsafe or corrupted: {str(excel_error)}"
                    )
                else:
                    # Non-XML related error, re-raise
                    logger.error(f"T12 Excel file parsing failed (non-XML error): {file_path}. Error: {excel_error}")
                    raise excel_error

            extracted_data = {
                "file_type": "excel",
                "document_type": "t12",
                "sheets": [],
                "total_sheets": len(workbook.sheetnames)
            }

            # Extract data from each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                print(f"    Processing sheet: {sheet_name}")

                # Get all data from the sheet using memory-efficient iteration
                sheet_data = iter_rows_efficiently(sheet, max_consecutive_empty_rows=20)
                print(f"    Sheet {sheet_name}: {len(sheet_data)} rows after efficient processing")

                # Clean the sheet data
                cleaned_sheet_data = clean_excel_data(sheet_data)
                print(f"    Sheet {sheet_name}: {len(cleaned_sheet_data)} rows after cleaning")

                sheet_info = {
                    "sheet_name": sheet_name,
                    "data": cleaned_sheet_data,
                    "rows": len(cleaned_sheet_data),
                    "columns": len(cleaned_sheet_data[0]) if cleaned_sheet_data else 0
                }

                extracted_data["sheets"].append(sheet_info)

            workbook.close()
            print(f"    T12 Excel extraction completed: {len(extracted_data['sheets'])} sheets")
            logger.info(f"T12 Excel extraction completed successfully: {file_path}")

            return extracted_data

        except HTTPException:
            # Re-raise HTTP exceptions as-is
            raise
        except Exception as e:
            print(f"    T12 Excel extraction error: {str(e)}")
            logger.error(f"T12 Excel extraction failed: {file_path}. Error: {str(e)}")
            raise HTTPException(
                status_code=500,
                detail=f"Failed to extract T12 Excel data: {str(e)}"
            )
