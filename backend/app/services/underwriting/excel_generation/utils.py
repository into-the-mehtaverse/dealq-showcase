"""
Excel utility functions for DealQ Excel Generation Service.

This module provides helper functions for working with Excel files using openpyxl.
Moved from app/utils/excel_utils.py as part of service modularization.
"""

from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string


def write_data_with_mapping(
    worksheet: Worksheet,
    data: List[Dict[str, Any]],
    mapping_config: Dict[str, Any]
) -> None:
    """
    Write data to worksheet using column mapping configuration.

    Args:
        worksheet: The openpyxl worksheet to write to
        data: List of dictionaries containing the data
        mapping_config: Mapping configuration with columns and data_start_row
    """
    columns = mapping_config.get("columns", {})
    data_start_row = mapping_config.get("data_start_row", 1)

    # Clear existing data first
    if data:
        # Get column range for clearing
        column_letters = [col for col in columns.values() if col]
        if column_letters:
            column_numbers = [column_index_from_string(col) for col in column_letters]
            start_col = min(column_numbers)
            end_col = max(column_numbers)

            clear_data_rows(
                worksheet=worksheet,
                start_row=data_start_row,
                end_row=data_start_row + len(data) + 10,  # Clear a bit extra for safety
                start_column=start_col,
                end_column=end_col
            )

    # Write data using column mappings
    for row_index, row_data in enumerate(data):
        current_row = data_start_row + row_index

        for field_name, column_letter in columns.items():
            if column_letter and field_name in row_data:
                column_number = column_index_from_string(column_letter)
                value = row_data[field_name]
                worksheet.cell(row=current_row, column=column_number, value=value)


def write_property_info_with_mapping(
    workbook,
    property_data: Dict[str, Any],
    mapping_config: Dict[str, Any]
) -> None:
    """
    Write property information to specific cells using single-cell mapping configuration.

    Args:
        workbook: The openpyxl workbook to write to
        property_data: Dictionary containing property information
        mapping_config: Mapping configuration with cell locations for each field
    """
    if mapping_config.get("type") != "property_info":
        return

    cells = mapping_config.get("cells", {})

    for field_name, cell_info in cells.items():
        if field_name in property_data and cell_info:
            sheet_name = cell_info.get("sheet")
            cell_address = cell_info.get("cell")

            if sheet_name and cell_address and sheet_name in workbook.sheetnames:
                try:
                    worksheet = workbook[sheet_name]
                    value = property_data[field_name]
                    worksheet[cell_address] = value
                except Exception as e:
                    print(f"Warning: Failed to write {field_name} to {sheet_name}!{cell_address}: {str(e)}")
                    continue


def clear_data_rows(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    start_column: int = 1,
    end_column: int = 10
) -> None:
    """
    Clear data in a range of rows and columns, preserving formulas.

    Args:
        worksheet: The openpyxl worksheet to clear
        start_row: First row to clear (1-indexed)
        end_row: Last row to clear (1-indexed)
        start_column: First column to clear (1-indexed)
        end_column: Last column to clear (1-indexed)
    """
    for row in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            cell = worksheet.cell(row=row, column=col)
            # Only clear if it's not a formula
            if cell.value is not None and not str(cell.value).startswith('='):
                cell.value = None
