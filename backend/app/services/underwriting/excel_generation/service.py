"""
Unified Excel Generation Service for Real Estate Document Analysis

Focuses solely on Excel generation logic without storage concerns.
Storage operations are handled by the StorageService and orchestrated by pipeline stages.
"""

import os
from typing import Dict, List, Any, Optional
from pathlib import Path
from fastapi import HTTPException
import openpyxl
from openpyxl.workbook import Workbook

from .utils import write_data_with_mapping, write_property_info_with_mapping


class ExcelGenerationService:
    """Service focused on Excel generation logic only."""

    def __init__(self):
        """Initialize the Excel generation service."""
        pass

    def fill_combined_data(
        self,
        structured_data: Dict[str, Any],
        model_mapping: Dict[str, Any],
        template_path: str,
        output_path: str = None
    ) -> str:
        """
        Fill both rent roll and T12 data into the Excel template using dynamic mappings.

        Args:
            structured_data: Dictionary containing both rent_roll and t12 data
            model_mapping: Complete model mapping configuration
            template_path: Path to the Excel template file
            output_path: Path where the filled Excel file should be saved (optional)

        Returns:
            str: Path to the saved Excel file
        """
        try:
            # Validate template exists
            if not os.path.exists(template_path):
                raise HTTPException(
                    status_code=404,
                    detail=f"Excel template not found at: {template_path}"
                )

            # Generate output path if not provided
            if not output_path:
                import uuid
                unique_id = str(uuid.uuid4())
                output_dir = "files/outputs"
                os.makedirs(output_dir, exist_ok=True)
                output_path = os.path.join(output_dir, f"dealq-{unique_id}.xlsm")

            # Load the Excel template with VBA macros preserved
            workbook = openpyxl.load_workbook(template_path, keep_vba=True)

            # Process rent roll data if available
            if "rent_roll" in structured_data and "rent_roll" in model_mapping.get("mappings", {}):
                rent_roll_data = structured_data["rent_roll"]
                rent_roll_mapping = model_mapping["mappings"]["rent_roll"]

                if rent_roll_data and rent_roll_mapping:
                    sheet_name = rent_roll_mapping.get("sheet_name")
                    if sheet_name and sheet_name in workbook.sheetnames:
                        # Process rent roll using dynamic column mapping
                        rr_sheet = workbook[sheet_name]

                        # Use the dynamic mapping function
                        write_data_with_mapping(
                            worksheet=rr_sheet,
                            data=rent_roll_data,
                            mapping_config=rent_roll_mapping
                        )

            # Process T12 data if available
            if "t12" in structured_data and "t12" in model_mapping.get("mappings", {}):
                t12_data = structured_data["t12"]
                t12_mapping = model_mapping["mappings"]["t12"]

                if t12_data and t12_mapping:
                    sheet_name = t12_mapping.get("sheet_name")
                    if sheet_name and sheet_name in workbook.sheetnames:
                        # Process T12 data using dynamic column mapping
                        t12_sheet = workbook[sheet_name]

                        # Use the dynamic mapping function
                        write_data_with_mapping(
                            worksheet=t12_sheet,
                            data=t12_data,
                            mapping_config=t12_mapping
                        )

            # Process property info if available
            if "property_info" in structured_data and "property_info" in model_mapping.get("mappings", {}):
                property_data = structured_data["property_info"]
                property_mapping = model_mapping["mappings"]["property_info"]

                if property_data and property_mapping:
                    # Process property info using single-cell mapping
                    write_property_info_with_mapping(
                        workbook=workbook,
                        property_data=property_data,
                        mapping_config=property_mapping
                    )

            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # Save the updated workbook
            workbook.save(output_path)
            workbook.close()

            return output_path

        except Exception as e:
            if isinstance(e, HTTPException):
                raise e
            raise HTTPException(
                status_code=500,
                detail=f"Error filling combined data in Excel template: {str(e)}"
            )

    def generate_excel(
        self,
        structured_data: Dict[str, Any],
        template_path: str,
        model_mapping: Dict[str, Any],
        output_filename: str = None
    ) -> str:
        """
        Generate Excel file from structured data using provided template and mapping.

        Args:
            structured_data: Structured data containing rent_roll and t12 data
            template_path: Path to the Excel template file
            model_mapping: Model mapping configuration
            output_filename: Optional custom output filename

        Returns:
            Path to the generated Excel file
        """
        try:
            # Generate unique output filename if not provided
            if not output_filename:
                import uuid
                unique_id = str(uuid.uuid4())
                output_filename = f"dealq-{unique_id}.xlsm"

            # Ensure output directory exists
            output_dir = "files/outputs"
            os.makedirs(output_dir, exist_ok=True)

            output_path = os.path.join(output_dir, output_filename)

            # Use the existing Excel filler service to generate the file
            result_path = self.fill_combined_data(
                structured_data=structured_data,
                model_mapping=model_mapping,
                template_path=template_path,
                output_path=output_path
            )

            return result_path

        except Exception as e:
            if isinstance(e, HTTPException):
                raise e
            raise HTTPException(
                status_code=500,
                detail=f"Excel generation failed: {str(e)}"
            )

    def get_template_info(self, template_path: str) -> Dict[str, Any]:
        """
        Get information about the Excel template structure.

        Args:
            template_path: Path to the Excel template file

        Returns:
            Dict with template information including sheet names and RR sheet structure
        """
        try:
            if not os.path.exists(template_path):
                return {"error": f"Template not found at: {template_path}"}

            workbook = openpyxl.load_workbook(template_path)

            info = {
                "template_path": template_path,
                "sheet_names": workbook.sheetnames,
                "rr_sheet_exists": 'RR' in workbook.sheetnames
            }

            if 'RR' in workbook.sheetnames:
                rr_sheet = workbook['RR']
                info["rr_sheet_info"] = {
                    "max_row": rr_sheet.max_row,
                    "max_column": rr_sheet.max_column,
                    "headers_row_10": [
                        rr_sheet.cell(row=10, column=col).value
                        for col in range(3, 9)  # Columns C through H
                    ]
                }

            workbook.close()
            return info

        except Exception as e:
            return {"error": f"Could not read template: {str(e)}"}

    def validate_structured_data(self, structured_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Validate that structured data has the expected format for Excel generation.

        Args:
            structured_data: Data to validate

        Returns:
            Dictionary with validation results
        """
        validation_result = {
            "valid": True,
            "errors": [],
            "warnings": [],
            "data_summary": {}
        }

        try:
            # Check for expected data types
            if not isinstance(structured_data, dict):
                validation_result["valid"] = False
                validation_result["errors"].append("Structured data must be a dictionary")
                return validation_result

            # Check rent roll data
            if "rent_roll" in structured_data:
                rent_roll_data = structured_data["rent_roll"]
                if isinstance(rent_roll_data, list):
                    validation_result["data_summary"]["rent_roll_units"] = len(rent_roll_data)
                    if len(rent_roll_data) == 0:
                        validation_result["warnings"].append("No rent roll data found")
                else:
                    validation_result["errors"].append("Rent roll data must be a list")
                    validation_result["valid"] = False
            else:
                validation_result["warnings"].append("No rent roll data provided")

            # Check T12 data
            if "t12" in structured_data:
                t12_data = structured_data["t12"]
                if isinstance(t12_data, list):
                    validation_result["data_summary"]["t12_line_items"] = len(t12_data)
                    if len(t12_data) == 0:
                        validation_result["warnings"].append("No T12 data found")
                else:
                    validation_result["errors"].append("T12 data must be a list")
                    validation_result["valid"] = False
            else:
                validation_result["warnings"].append("No T12 data provided")

            # Check property info data
            if "property_info" in structured_data:
                property_data = structured_data["property_info"]
                if isinstance(property_data, dict):
                    validation_result["data_summary"]["property_info_fields"] = len(property_data)
                    if len(property_data) == 0:
                        validation_result["warnings"].append("No property info data found")
                else:
                    validation_result["errors"].append("Property info data must be a dictionary")
                    validation_result["valid"] = False
            else:
                validation_result["warnings"].append("No property info data provided")

            return validation_result

        except Exception as e:
            validation_result["valid"] = False
            validation_result["errors"].append(f"Validation error: {str(e)}")
            return validation_result


# Service instance is created in core/dependencies/services.py
